'''
File Name: autoExcel.py

Author: hxoreyer

Version: 1.0

Date: 2020/12/18

Description:
    用于对比两个Excel文件的编号和价格，默认文件为
    src.xlsx,需要修改的文件自行拖拽，最终将拖拽文件
    的价格修改为默认文件的价格，并生成新Excel文件

'''


import openpyxl as opx
import sys

def readexcel(name): #获得Sheet
    wb = opx.load_workbook(name)
    ws = wb['Sheet1']
    return wb,ws

def getPriceCol(ws): #获得价格在第几行
    pricecol = 0
    for col in range(1,100):
        if ws.cell(1,col).value == None:
            break
        if ws.cell(1,col).value == '价格':
            pricecol = col
            break
    #print(pricecol)
    return pricecol

def getCodeNum(ws): #获得条形码编号
    codenum = 0
    for col in range(1,10):
        if ws.cell(1,col).value == None:
            break
        if ws.cell(1,col).value == '编号':
            codenum = col
            break
    #print(codenum)
    return codenum

def getSelfExcel(): #获得系统excel的sheet
    return readexcel('src.xlsx')

def getFileExcel(name): #获得文件sheet
    return readexcel(name)

def getDict(ws): #给库数据写入字典
    dc = {}
    ncode = getCodeNum(ws)
    nprice = getPriceCol(ws)
    for row in range(2,sys.maxsize):
        if ws.cell(row,ncode).value == None:
            break
        dc[ws.cell(row,ncode).value] = ws.cell(row,nprice).value
    return dc

def doJob(ws,dc): #完成价格替换
    ncode = getCodeNum(ws)
    nprice = getPriceCol(ws)
    for row in range(2,sys.maxsize):
        if ws.cell(row,ncode).value == None:
            break
        ws.cell(row,nprice,dc[ws.cell(row,ncode).value])

def main(): #主函数
    _,ws1 = getSelfExcel()
    dc = getDict(ws1)
    wb2,ws2 = getFileExcel(sys.argv[1])
    doJob(ws2,dc)
    wb2.save('结果.xlsx')
    print('保存成功，输入回车键关闭窗口!')
    input()

if __name__ == '__main__':
    main()